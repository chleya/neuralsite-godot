from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import xlrd


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "converted-import-package"
DESIGN_VERSION = "351-import-v1"
SKIP_FILES = {"400章桥梁汇总.xlsm"}


@dataclass
class WorkAreaRecord:
    id: str
    name: str
    type: str
    owner: str
    planned_progress: float
    actual_progress: float
    description: str


@dataclass
class DesignQuantityRecord:
    id: str
    work_area_id: str
    item_name: str
    item_code: str
    category: str
    unit: str
    target_quantity: float
    design_version: str
    notes: str


def main() -> None:
    excel_files = [
        path
        for path in sorted(ROOT.rglob("*"))
        if path.is_file() and path.suffix.lower() in {".xls", ".xlsx", ".xlsm"} and path.name not in SKIP_FILES
    ]

    work_areas: list[WorkAreaRecord] = []
    design_quantities: list[DesignQuantityRecord] = []
    manifest: list[dict] = []

    for index, path in enumerate(excel_files, start=1):
        work_area_id = f"w351_{index:03d}"
        work_area_name = path.stem
        work_area_type = infer_work_area_type(path)
        sheet_name, quantity_column, rows = extract_sheet_rows(path)
        if not rows:
            manifest.append(
                {
                    "file": str(path.relative_to(ROOT)),
                    "status": "skipped",
                    "reason": "No quantifiable rows found",
                }
            )
            continue

        work_areas.append(
            WorkAreaRecord(
                id=work_area_id,
                name=work_area_name,
                type=work_area_type,
                owner="",
                planned_progress=0.0,
                actual_progress=0.0,
                description=f"Imported from 351 quantity list: {path.relative_to(ROOT)}",
            )
        )

        for row_idx, row in enumerate(rows, start=1):
            design_quantities.append(
                DesignQuantityRecord(
                    id="",
                    work_area_id=work_area_id,
                    item_name=row["item_name"],
                    item_code=row["item_code"],
                    category=infer_category(path),
                    unit=row["unit"],
                    target_quantity=row["quantity"],
                    design_version=DESIGN_VERSION,
                    notes=f"source_file={path.relative_to(ROOT)}; sheet={sheet_name}; source_column={quantity_column}; row={row['source_row']}",
                )
            )

        manifest.append(
            {
                "file": str(path.relative_to(ROOT)),
                "work_area_id": work_area_id,
                "work_area_name": work_area_name,
                "sheet": sheet_name,
                "quantity_column": quantity_column,
                "row_count": len(rows),
                "status": "ok",
            }
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_work_areas(work_areas, OUTPUT_DIR / "work_areas.csv")
    write_design_quantities(design_quantities, OUTPUT_DIR / "design_quantities.csv")
    (OUTPUT_DIR / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "README.md").write_text(build_readme(work_areas, design_quantities, manifest), encoding="utf-8")

    print(f"Generated {len(work_areas)} work areas and {len(design_quantities)} design quantities in {OUTPUT_DIR}")


def extract_sheet_rows(path: Path) -> tuple[str, str, list[dict]]:
    workbook, sheets = load_workbook_rows(path)
    preferred = next((name for name in sheets if "工程量清单" in name), sheets[0])
    rows = workbook[preferred]
    return parse_rows(preferred, rows)


def load_workbook_rows(path: Path) -> tuple[dict[str, list[list[object]]], list[str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        data = {
            name: [list(row) for row in wb[name].iter_rows(values_only=True)]
            for name in sheet_names
        }
        return data, sheet_names

    wb = xlrd.open_workbook(path)
    sheet_names = wb.sheet_names()
    data = {}
    for name in sheet_names:
        sh = wb.sheet_by_name(name)
        data[name] = [sh.row_values(i) for i in range(sh.nrows)]
    return data, sheet_names


def parse_rows(sheet_name: str, rows: list[list[object]]) -> tuple[str, str, list[dict]]:
    header_index = None
    header_map: dict[str, int] = {}

    for idx, row in enumerate(rows):
        normalized = [normalize_text(cell) for cell in row]
        if "子目号" in normalized and any("子目名称" in value or "子  目  名  称" in value for value in normalized):
            header_index = idx
            for col_idx, value in enumerate(normalized):
                header_map[value] = col_idx
            break

    if header_index is None:
        return sheet_name, "", []

    quantity_column_name = ""
    for candidate in ("核算数量", "图纸数量", "清单数量", "数量"):
        for header_text in header_map:
            if candidate in header_text:
                quantity_column_name = header_text
                break
        if quantity_column_name:
            break

    if not quantity_column_name:
        return sheet_name, "", []

    code_idx = find_header_index(header_map, "子目号")
    name_idx = find_header_index(header_map, "子目名称") or find_header_index(header_map, "子  目  名  称")
    unit_idx = find_header_index(header_map, "单位")
    qty_idx = header_map[quantity_column_name]
    if code_idx is None or name_idx is None or unit_idx is None:
        return sheet_name, quantity_column_name, []

    extracted: list[dict] = []
    for source_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        item_code = normalize_code(get_cell(row, code_idx))
        item_name = normalize_text(get_cell(row, name_idx))
        unit = normalize_text(get_cell(row, unit_idx))
        quantity = parse_number(get_cell(row, qty_idx))
        if not item_code or not item_name or not unit or quantity is None:
            continue
        extracted.append(
            {
                "item_code": item_code,
                "item_name": item_name,
                "unit": unit,
                "quantity": quantity,
                "source_row": source_row,
            }
        )

    return sheet_name, quantity_column_name, extracted


def infer_work_area_type(path: Path) -> str:
    name = path.stem
    if "桥" in name or "钢管" in name:
        return "bridge"
    if "路" in name:
        return "road"
    return "general"


def infer_category(path: Path) -> str:
    name = path.name
    if "200章" in name or "路基" in name:
        return "earthwork"
    if "300章" in name or "路面" in name:
        return "paving"
    if "桥" in name or "400章桥梁" in name or "钢管" in name:
        return "bridge"
    if "涵洞" in name:
        return "culvert"
    if "交通安全" in name or "600章" in name:
        return "traffic_safety"
    if "绿化" in name or "700" in name or "环境保护" in name:
        return "greening"
    return "general"


def write_work_areas(rows: Iterable[WorkAreaRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "name", "type", "owner", "planned_progress", "actual_progress", "description"])
        for row in rows:
            writer.writerow([row.id, row.name, row.type, row.owner, row.planned_progress, row.actual_progress, row.description])


def write_design_quantities(rows: Iterable[DesignQuantityRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "work_area_id", "item_name", "item_code", "category", "unit", "target_quantity", "design_version", "notes"])
        for row in rows:
            writer.writerow([row.id, row.work_area_id, row.item_name, row.item_code, row.category, row.unit, row.target_quantity, row.design_version, row.notes])


def build_readme(work_areas: list[WorkAreaRecord], design_quantities: list[DesignQuantityRecord], manifest: list[dict]) -> str:
    ok_files = [item for item in manifest if item["status"] == "ok"]
    skipped_files = [item for item in manifest if item["status"] != "ok"]
    return "\n".join(
        [
            "# 351 清单抽取结果",
            "",
            "这批文件由 `extract_design_quantities.py` 从真实 Excel 清单自动抽取生成。",
            "",
            f"- 工作面数量: {len(work_areas)}",
            f"- 设计工程量条目数: {len(design_quantities)}",
            f"- 成功抽取文件数: {len(ok_files)}",
            f"- 跳过文件数: {len(skipped_files)}",
            "",
            "输出文件：",
            "- `work_areas.csv`",
            "- `design_quantities.csv`",
            "- `manifest.json`",
            "",
            "说明：",
            "- 当前只抽取适合映射到 `design_quantities` 的目标量数据",
            "- 默认优先取 `核算数量`，其次 `图纸数量`，再其次 `清单数量/数量`",
            "- `400章桥梁汇总.xlsm` 默认跳过，避免与分桥清单重复统计",
            "- `notes` 字段会保留来源文件、工作表和原始行号",
        ]
    )


def find_header_index(header_map: dict[str, int], target: str) -> int | None:
    for header, idx in header_map.items():
        if target in header:
            return idx
    return None


def get_cell(row: list[object], index: int) -> object:
    return row[index] if index < len(row) else None


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_code(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if not re.search(r"\d", text):
        return ""
    return text


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


if __name__ == "__main__":
    main()
